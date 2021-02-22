import os

venv = os.path.join(os.getcwd(), "cadets_duties", "bin", "activate_this.py")
exec(open(venv).read(), {'__file__': venv})

import random
import openpyxl
from openpyxl import Workbook

try:
    fname = input("Class List Filename: ")
    if fname.endswith(".xls"):
        fname += "x"
    elif not fname.endswith(".xlsx"):
        fname += ".xlsx"
    else: pass
    wb = openpyxl.load_workbook(fname)
except:
    raise FileNotFoundError(f"'{fname}' file does not exist.")
    quit()
nwb = Workbook()
ws1 = nwb.active
nwb.remove(ws1)

ws = wb[wb.sheetnames[0]]

def randomize_duties(class_list, class_count):
    """Receives a list of students and randomly assigns each student to a duty post.
    """

    #Create tables for each duty post
    time = ["MORNING", "LUNCH", "RECESS"]
    places = ["PANAY", "ICB", "QUADRANGLE"]
    AB = ("A","B")
    sheets = [nwb.create_sheet(places[place], place) for place in range(len(places))]
    #Separate boys from girls
    boys = []
    girls = []
    for student in class_list:
        if student[0].startswith('M'):
            boys.append(student)
        else:
            girls.append(student)
    genders = [boys, girls]
    #Append each student in a duty post
    for gender in genders:
        places_count = [0, 0, 0] #Create count for each place
        #Create a list for counting students
        gender_range = len(gender)
        num = [x for x in range(len(gender))]
        while len(num) > 0: #While there is a student in the count_list
            for place in range(len(places)): #iterates through each places
                sheet = nwb[places[place]]
                dimension = sheet.calculate_dimension()
                letter = dimension[3]
                if gender is girls: y = ord(letter)+1
                else: y = ord(letter)
                while places_count[place] < gender_range/3:
                    m = random.choice(num)
                    for stdnt in gender:
                        if (int(stdnt[0][1:])-1) == m:
                            #append it to the workbook
                            for x in range(2):
                                print(stdnt[x])
                                sheet[f"{chr(y)}{x+1}"] = stdnt[x]
                            y += 1
                            gender.remove(stdnt)
                            places_count[place] += 1
                    num.remove(m)
                    if len(num) < 1: break
    filename = input("Save as: ")
    if not filename.endswith(".xlsx"): filename += ".xlsx"
    nwb.save(filename)

class_list = []
for row in ws.values:
    student = []
    for cell in row:
        student.append(cell)
    if student[0] == "CN": continue
    class_list.append(tuple(student))
class_count = len(class_list)
randomize_duties(class_list, class_count)
